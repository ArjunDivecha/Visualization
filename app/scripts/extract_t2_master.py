#!/usr/bin/env python3
"""Extract T2 Master workbook into chart-friendly JSON for the web app."""

from __future__ import annotations

import argparse
import json
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def normalize_date(value: Any) -> str | None:
    """Convert Excel date-like values to YYYY-MM-DD strings."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    return str(value)


def to_float(value: Any) -> float | None:
    """Convert numeric values to float; return None for blanks/non-numeric."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def extract_workbook(input_xlsx: Path) -> dict[str, Any]:
    wb = load_workbook(input_xlsx, data_only=True, read_only=True)
    output: dict[str, Any] = {
        "generated_at": datetime.now(UTC).isoformat(timespec="seconds"),
        "source_file": str(input_xlsx),
        "sheets": {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2 or ws.max_column < 2:
            continue

        row_iter = ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
            values_only=True,
        )

        try:
            header_row = next(row_iter)
        except StopIteration:
            continue

        headers: list[str | None] = []
        for raw in header_row[1:]:
            if raw is None or str(raw).strip() == "":
                headers.append(None)
            else:
                headers.append(str(raw).strip())

        countries = [h for h in headers if h]
        rows = []

        for raw_row in row_iter:
            dt = normalize_date(raw_row[0] if raw_row else None)
            if not dt:
                continue

            values: dict[str, float | None] = {}
            has_any_value = False
            for header, raw_value in zip(headers, raw_row[1:]):
                if not header:
                    continue
                num = to_float(raw_value)
                values[header] = num
                if num is not None:
                    has_any_value = True

            if has_any_value:
                rows.append({"date": dt, "values": values})

        if rows and countries:
            output["sheets"][sheet_name] = {
                "countries": countries,
                "rows": rows,
            }

    wb.close()
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract workbook to JSON")
    parser.add_argument("--input", required=True, help="Path to input .xlsx")
    parser.add_argument("--output", required=True, help="Path to output .json")
    args = parser.parse_args()

    input_xlsx = Path(args.input).expanduser().resolve()
    output_json = Path(args.output).expanduser().resolve()
    output_json.parent.mkdir(parents=True, exist_ok=True)

    payload = extract_workbook(input_xlsx)
    output_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {output_json}")
    print(f"Sheets exported: {len(payload['sheets'])}")


if __name__ == "__main__":
    main()
